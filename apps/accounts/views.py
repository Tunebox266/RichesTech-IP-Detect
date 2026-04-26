# accounts/views.py
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import login, authenticate, logout, update_session_auth_hash
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib import messages
from django.db.models import Q, Count, Sum
from django.core.paginator import Paginator
from django.http import JsonResponse, HttpResponse, FileResponse
from django.utils import timezone
from django.views.decorators.http import require_POST
from django.contrib.auth.tokens import default_token_generator
from django.utils.http import urlsafe_base64_encode, urlsafe_base64_decode
from django.utils.encoding import force_bytes, force_str
from django.core.mail import send_mail, EmailMessage
from django.template.loader import render_to_string
from django.conf import settings
from django.db import transaction
import csv
import openpyxl
import qrcode
from io import BytesIO
from datetime import timedelta
import json

from .models import (
    User, ActivityLog, LoginAttempt, StudentExecutive,
    ExecutiveMeeting, ExecutiveTask, ExecutiveDiscussion,
    DiscussionComment, MeetingAttendance
)
from .forms import (
    # Authentication forms
    CustomAuthenticationForm,
    CustomPasswordChangeForm,
    CustomPasswordResetForm,
    CustomSetPasswordForm,
    
    # User management forms
    CustomUserCreationForm,
    CustomUserChangeForm,
    StudentProfileForm,
    StaffProfileForm,
    ProfileImageForm,
    BulkStudentUploadForm,
    UserSearchForm,
    
    # Executive forms
    StudentExecutiveForm,
    ExecutiveMeetingForm,
    ExecutiveTaskForm,
    ExecutiveTaskUpdateForm,
    ExecutiveDiscussionForm,
    DiscussionCommentForm,
    MeetingAttendanceForm,
)

from apps.core.models import AcademicSetting
from apps.courses.models import Course, CourseRegistration
from apps.payments.models import Payment
from apps.core.decorators import passcode_required

# ========== HELPER FUNCTIONS ==========

def get_client_ip(request):
    """Get client IP address from request"""
    x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
    if x_forwarded_for:
        ip = x_forwarded_for.split(',')[0]
    else:
        ip = request.META.get('REMOTE_ADDR')
    return ip


def log_activity(user, action_type, request, details=None):
    """Helper function to log user activities"""
    ActivityLog.objects.create(
        user=user,
        action_type=action_type,
        ip_address=get_client_ip(request),
        details=details or {}
    )


def send_password_email(user, temp_password):
    """Send password email to new user"""
    subject = 'Your MELTSA Account Has Been Created'
    message = render_to_string('accounts/email/new_account.html', {
        'user': user,
        'temp_password': temp_password,
        'login_url': settings.SITE_URL + '/accounts/login/'
    })
    send_mail(
        subject,
        message,
        settings.DEFAULT_FROM_EMAIL,
        [user.email],
        fail_silently=False,
    )


def check_account_lock(username):
    """Check if account is locked"""
    try:
        user = User.objects.get(
            Q(student_id=username) | 
            Q(staff_id=username) | 
            Q(username=username)
        )
        
        if user.account_locked_until and user.account_locked_until > timezone.now():
            return True, user.account_locked_until
        return False, None
    except User.DoesNotExist:
        return False, None


def handle_failed_login(username, request):
    """Handle failed login attempts"""
    try:
        user = User.objects.get(
            Q(student_id=username) | 
            Q(staff_id=username) | 
            Q(username=username)
        )
        user.failed_login_attempts += 1
        
        if user.failed_login_attempts >= 5:
            user.account_locked_until = timezone.now() + timedelta(minutes=15)
        
        user.save()
        
        # Log failed attempt
        LoginAttempt.objects.create(
            username=username,
            ip_address=get_client_ip(request),
            successful=False
        )
    except User.DoesNotExist:
        # Log attempt for non-existent user
        LoginAttempt.objects.create(
            username=username,
            ip_address=get_client_ip(request),
            successful=False
        )


# ========== AUTHENTICATION VIEWS ==========
@passcode_required
def login_view(request):
    """Handle user login"""
    if request.user.is_authenticated:
        return redirect('accounts:dashboard_redirect')
    
    if request.method == 'POST':
        form = CustomAuthenticationForm(request, data=request.POST)
        if form.is_valid():
            username = form.cleaned_data.get('username')
            password = form.cleaned_data.get('password')
            remember_me = form.cleaned_data.get('remember_me', False)
            
            # Check if account is locked
            is_locked, lock_until = check_account_lock(username)
            if is_locked:
                messages.error(
                    request, 
                    f'Account locked until {lock_until.strftime("%Y-%m-%d %H:%M")} due to multiple failed attempts.'
                )
                return render(request, 'accounts/login.html', {'form': form})
            
            # Authenticate
            user = authenticate(request, username=username, password=password)
            
            if user is not None:
                if user.is_active:
                    # Reset failed attempts
                    user.failed_login_attempts = 0
                    user.account_locked_until = None
                    user.save()
                    
                      # Clear passcode session flags AFTER successful login
                    if 'passcode_verified' in request.session:
                        del request.session['passcode_verified']
                    if 'passcode_verified_time' in request.session:
                        del request.session['passcode_verified_time']
                    
                    # Log successful login
                    log_activity(user, 'login', request, {
                        'username': username,
                        'user_agent': request.META.get('HTTP_USER_AGENT', '')
                    })
                    
                    LoginAttempt.objects.create(
                        username=username,
                        ip_address=get_client_ip(request),
                        successful=True
                    )
                    
                    login(request, user)
                    
                    # Set session expiry based on remember_me
                    if not remember_me:
                        request.session.set_expiry(0)  # Browser close
                    
                    # Check if password change required
                    if user.requires_password_change:
                        messages.warning(
                            request, 
                            'This is your first login. Please change your password for security.'
                        )
                        return redirect('accounts:password_change')
                    
                    messages.success(request, f'Welcome back, {user.get_full_name()}!')
                    return redirect('accounts:dashboard_redirect')
                else:
                    messages.error(request, 'Your account has been deactivated. Please contact administrator.')
            else:
                handle_failed_login(username, request)
                messages.error(request, 'Invalid username or password.')
    else:
        form = CustomAuthenticationForm()
    
    return render(request, 'accounts/login.html', {'form': form})


def logout_view(request):
    """Handle user logout"""
    if request.user.is_authenticated:
        log_activity(request.user, 'logout', request)
        logout(request)
        messages.success(request, 'You have been logged out successfully.')
    
    return redirect('core:home')


@login_required
def password_change(request):
    """Change user password"""
    if request.method == 'POST':
        form = CustomPasswordChangeForm(request.user, request.POST, request=request)
        if form.is_valid():
            user = form.save()
            update_session_auth_hash(request, user)
            
            # Update requires_password_change flag
            user.requires_password_change = False
            user.save()
            
            # Log activity
            log_activity(user, 'password_change', request)
            
            messages.success(request, 'Your password was changed successfully!')
            return redirect('accounts:profile')
    else:
        form = CustomPasswordChangeForm(request.user)
        form.request = request
    
    return render(request, 'accounts/password_change.html', {'form': form})


def password_reset_request(request):
    """Request password reset"""
    if request.method == 'POST':
        form = CustomPasswordResetForm(request.POST)
        if form.is_valid():
            email = form.cleaned_data['email']
            users = User.objects.filter(email=email)
            
            if users.exists():
                for user in users:
                    # Generate reset token
                    token = default_token_generator.make_token(user)
                    uid = urlsafe_base64_encode(force_bytes(user.pk))
                    
                    # Build reset URL
                    reset_url = request.build_absolute_uri(
                        f'/accounts/password-reset/{uid}/{token}/'
                    )
                    
                    # Send email
                    subject = 'Password Reset Request - MELTSA-TaTU'
                    message = render_to_string('accounts/email/password_reset.html', {
                        'user': user,
                        'reset_url': reset_url,
                        'site_name': 'MELTSA-TaTU',
                        'valid_hours': 24
                    })
                    
                    send_mail(
                        subject,
                        message,
                        settings.DEFAULT_FROM_EMAIL,
                        [user.email],
                        fail_silently=False,
                    )
                
                messages.success(
                    request, 
                    'Password reset instructions have been sent to your email. The link expires in 24 hours.'
                )
                return redirect('accounts:password_reset_done')
            else:
                messages.error(request, 'No user found with this email address.')
    else:
        form = CustomPasswordResetForm()
    
    return render(request, 'accounts/password_reset.html', {'form': form})


def password_reset_confirm(request, uidb64, token):
    """Confirm password reset"""
    try:
        uid = force_str(urlsafe_base64_decode(uidb64))
        user = User.objects.get(pk=uid)
    except (TypeError, ValueError, OverflowError, User.DoesNotExist):
        user = None
    
    if user is not None and default_token_generator.check_token(user, token):
        if request.method == 'POST':
            form = CustomSetPasswordForm(user, request.POST)
            if form.is_valid():
                user = form.save()
                user.requires_password_change = False
                user.save()
                
                log_activity(user, 'password_change', request, {'method': 'reset'})
                
                messages.success(request, 'Your password has been reset successfully! You can now login.')
                return redirect('accounts:password_reset_complete')
        else:
            form = CustomSetPasswordForm(user)
        
        return render(request, 'accounts/password_reset_confirm.html', {'form': form})
    else:
        messages.error(request, 'The password reset link is invalid or has expired.')
        return redirect('accounts:password_reset')


def password_reset_done(request):
    """Password reset request done"""
    return render(request, 'accounts/password_reset_done.html')


def password_reset_complete(request):
    """Password reset complete"""
    return render(request, 'accounts/password_reset_complete.html')


# ========== PROFILE VIEWS ==========

@login_required
def profile_view(request):
    """View user profile"""
    # Get user statistics based on user type
    context = {'active_tab': 'overview'}
    
    if request.user.user_type in ['student', 'executive']:
        # Get current academic setting
        current_academic = AcademicSetting.objects.filter(is_active=True).first()
        
        # Get registered courses
        registered_courses = CourseRegistration.objects.filter(
            student=request.user,
            academic_setting=current_academic
        ).select_related('course').count() if current_academic else 0
        
        # Get payment information
        total_paid = Payment.objects.filter(
            student=request.user,
            status='success'
        ).aggregate(Sum('amount'))['amount__sum'] or 0
        
        payment_count = Payment.objects.filter(
            student=request.user,
            status='success'
        ).count()
        
        # Get recent payments
        recent_payments = Payment.objects.filter(
            student=request.user
        ).order_by('-created_at')[:5]
        
        context.update({
            'registered_courses': registered_courses,
            'total_paid': total_paid,
            'payment_count': payment_count,
            'recent_payments': recent_payments,
            'current_academic': current_academic,
        })
    
    if request.user.user_type == 'executive':
        # Get executive profile
        executive_profile = get_object_or_404(StudentExecutive, user=request.user)
        
        # Get executive statistics
        upcoming_meetings = ExecutiveMeeting.objects.filter(
            organized_by=executive_profile,
            date__gte=timezone.now().date(),
            status='scheduled'
        ).count()
        
        pending_tasks = ExecutiveTask.objects.filter(
            assigned_to=executive_profile,
            status__in=['pending', 'in_progress']
        ).count()
        
        recent_meetings = ExecutiveMeeting.objects.filter(
            organized_by=executive_profile
        ).order_by('-date')[:5]
        
        assigned_tasks = ExecutiveTask.objects.filter(
            assigned_to=executive_profile
        ).order_by('-due_date')[:5]
        
        context.update({
            'executive_profile': executive_profile,
            'upcoming_meetings': upcoming_meetings,
            'pending_tasks': pending_tasks,
            'recent_meetings': recent_meetings,
            'assigned_tasks': assigned_tasks,
        })
    
    elif request.user.user_type == 'staff':
        # Staff statistics
        courses_taught = Course.objects.filter(
            staff=request.user,
            is_active=True
        ).count()
        
        total_students = User.objects.filter(
            user_type='student',
            is_active=True
        ).count()
        
        context.update({
            'courses_taught': courses_taught,
            'total_students': total_students,
        })
    
    return render(request, 'accounts/profile.html', context)


@login_required
def profile_edit(request):
    """Edit user profile"""
    if request.user.user_type in ['student', 'executive']:
        form_class = StudentProfileForm
    else:
        form_class = StaffProfileForm
    
    if request.method == 'POST':
        form = form_class(request.POST, request.FILES, instance=request.user)
        if form.is_valid():
            form.save()
            messages.success(request, 'Your profile has been updated successfully!')
            return redirect('accounts:profile')
    else:
        form = form_class(instance=request.user)
    
    # Add additional forms for executives
    executive_form = None
    if request.user.user_type == 'executive':
        executive_profile = StudentExecutive.objects.get(user=request.user)
        if request.method == 'POST':
            executive_form = StudentExecutiveForm(
                request.POST, request.FILES, 
                instance=executive_profile
            )
            if executive_form.is_valid():
                executive_form.save()
        else:
            executive_form = StudentExecutiveForm(instance=executive_profile)
    
    return render(request, 'accounts/profile_edit.html', {
        'form': form,
        'executive_form': executive_form,
        'active_tab': 'edit'
    })


@login_required
@require_POST
def upload_profile_photo(request):
    """Upload profile photo via AJAX"""
    if 'photo' in request.FILES:
        photo = request.FILES['photo']
        
        # Validate file type
        if not photo.content_type.startswith('image/'):
            return JsonResponse({
                'success': False, 
                'error': 'File must be an image'
            })
        
        # Validate file size (max 2MB)
        if photo.size > 2 * 1024 * 1024:
            return JsonResponse({
                'success': False, 
                'error': 'Image size must be less than 2MB'
            })
        
        request.user.profile_image = photo
        request.user.save()
        
        log_activity(request.user, 'profile_update', request, {'action': 'photo_upload'})
        
        return JsonResponse({
            'success': True,
            'photo_url': request.user.profile_image.url
        })
    
    return JsonResponse({'success': False, 'error': 'No photo provided'})


# ========== DASHBOARD REDIRECT ==========

@login_required
def dashboard_redirect(request):
    """Redirect to appropriate dashboard based on user type"""
    if request.user.user_type == 'admin':
        return redirect('core:admin_dashboard')
    elif request.user.user_type == 'staff':
        return redirect('core:staff_dashboard')
    elif request.user.user_type == 'executive':
        return redirect('core:executive_dashboard')
    else:
        return redirect('core:student_dashboard')


# ========== EXECUTIVE MANAGEMENT VIEWS ==========

@login_required
@user_passes_test(lambda u: u.user_type == 'executive')
def executive_dashboard(request):
    """Executive dashboard view"""
    executive = get_object_or_404(StudentExecutive, user=request.user)
    
    # Get upcoming meetings
    upcoming_meetings = ExecutiveMeeting.objects.filter(
        participants=executive,
        date__gte=timezone.now().date(),
        status='scheduled'
    ).order_by('date', 'start_time')[:5]
    
    # Get tasks assigned to executive
    my_tasks = ExecutiveTask.objects.filter(
        assigned_to=executive
    ).exclude(status='completed').order_by('due_date')[:5]
    
    # Get recent discussions
    recent_discussions = ExecutiveDiscussion.objects.filter(
        meeting__participants=executive
    ).order_by('-created_at')[:5]
    
    # Get meetings organized by executive
    organized_meetings = ExecutiveMeeting.objects.filter(
        organized_by=executive,
        date__gte=timezone.now().date()
    ).order_by('date')[:3]
    
    # Get statistics
    stats = {
        'total_meetings': ExecutiveMeeting.objects.filter(
            participants=executive
        ).count(),
        'upcoming_count': upcoming_meetings.count(),
        'pending_tasks': my_tasks.count(),
        'attended_meetings': MeetingAttendance.objects.filter(
            executive=executive
        ).count(),
    }
    
    context = {
        'executive': executive,
        'upcoming_meetings': upcoming_meetings,
        'my_tasks': my_tasks,
        'recent_discussions': recent_discussions,
        'organized_meetings': organized_meetings,
        'stats': stats,
    }
    
    return render(request, 'accounts/executive/dashboard.html', context)


@login_required
@user_passes_test(lambda u: u.user_type == 'executive')
def meeting_list(request):
    """List all meetings"""
    executive = get_object_or_404(StudentExecutive, user=request.user)
    
    # Filter by status
    status = request.GET.get('status', 'upcoming')
    
    if status == 'upcoming':
        meetings = ExecutiveMeeting.objects.filter(
            participants=executive,
            date__gte=timezone.now().date(),
            status__in=['scheduled']
        )
    elif status == 'past':
        meetings = ExecutiveMeeting.objects.filter(
            participants=executive,
            date__lt=timezone.now().date()
        )
    elif status == 'organized':
        meetings = ExecutiveMeeting.objects.filter(
            organized_by=executive
        )
    else:
        meetings = ExecutiveMeeting.objects.filter(participants=executive)
    
    meetings = meetings.order_by('-date', '-start_time')
    
    # Pagination
    paginator = Paginator(meetings, 10)
    page = request.GET.get('page')
    meetings = paginator.get_page(page)
    
    context = {
        'meetings': meetings,
        'current_status': status,
        'executive': executive,
    }
    
    return render(request, 'accounts/executive/meeting_list.html', context)


@login_required
@user_passes_test(lambda u: u.user_type == 'executive')
def meeting_create(request):
    """Create new meeting"""
    executive = get_object_or_404(StudentExecutive, user=request.user)
    
    if request.method == 'POST':
        form = ExecutiveMeetingForm(request.POST, request.FILES)
        if form.is_valid():
            meeting = form.save(commit=False)
            meeting.organized_by = executive
            meeting.save()
            form.save_m2m()  # Save participants
            
            # Add organizer as participant if not already included
            if executive not in meeting.participants.all():
                meeting.participants.add(executive)
            
            log_activity(
                request.user, 
                'meeting_created', 
                request, 
                {'meeting_id': meeting.id, 'title': meeting.title}
            )
            
            messages.success(request, f'Meeting "{meeting.title}" created successfully!')
            return redirect('accounts:meeting_detail', pk=meeting.pk)
    else:
        form = ExecutiveMeetingForm()
    
    return render(request, 'accounts/executive/meeting_form.html', {
        'form': form,
        'action': 'Create'
    })


@login_required
@user_passes_test(lambda u: u.user_type == 'executive')
def meeting_detail(request, pk):
    """View meeting details"""
    meeting = get_object_or_404(ExecutiveMeeting, pk=pk)
    executive = get_object_or_404(StudentExecutive, user=request.user)
    
    # Check if user can view this meeting
    if executive not in meeting.participants.all() and meeting.organized_by != executive:
        messages.error(request, 'You do not have permission to view this meeting.')
        return redirect('accounts:meeting_list')
    
    # Get attendance
    attendance = MeetingAttendance.objects.filter(meeting=meeting).select_related('executive')
    
    # Check if user has checked in
    has_checked_in = attendance.filter(executive=executive).exists()
    
    # Get discussions
    discussions = ExecutiveDiscussion.objects.filter(meeting=meeting).order_by('-created_at')
    
    # Get tasks related to meeting
    tasks = ExecutiveTask.objects.filter(related_to_meeting=meeting)
    
    context = {
        'meeting': meeting,
        'attendance': attendance,
        'has_checked_in': has_checked_in,
        'discussions': discussions,
        'tasks': tasks,
        'executive': executive,
        'is_organizer': meeting.organized_by == executive,
    }
    
    return render(request, 'accounts/executive/meeting_detail.html', context)


@login_required
@user_passes_test(lambda u: u.user_type == 'executive')
def meeting_edit(request, pk):
    """Edit meeting"""
    meeting = get_object_or_404(ExecutiveMeeting, pk=pk)
    executive = get_object_or_404(StudentExecutive, user=request.user)
    
    # Only organizer can edit
    if meeting.organized_by != executive:
        messages.error(request, 'Only the meeting organizer can edit this meeting.')
        return redirect('accounts:meeting_detail', pk=meeting.pk)
    
    if request.method == 'POST':
        form = ExecutiveMeetingForm(request.POST, request.FILES, instance=meeting)
        if form.is_valid():
            form.save()
            messages.success(request, 'Meeting updated successfully!')
            return redirect('accounts:meeting_detail', pk=meeting.pk)
    else:
        form = ExecutiveMeetingForm(instance=meeting)
    
    return render(request, 'accounts/executive/meeting_form.html', {
        'form': form,
        'action': 'Edit',
        'meeting': meeting
    })


@login_required
@require_POST
@user_passes_test(lambda u: u.user_type == 'executive')
def meeting_check_in(request, pk):
    """Check in to meeting via QR code or code"""
    meeting = get_object_or_404(ExecutiveMeeting, pk=pk)
    executive = get_object_or_404(StudentExecutive, user=request.user)
    
    # Check if meeting is today
    if meeting.date != timezone.now().date():
        return JsonResponse({
            'success': False,
            'error': 'This meeting is not scheduled for today.'
        })
    
    # Check if already checked in
    if MeetingAttendance.objects.filter(meeting=meeting, executive=executive).exists():
        return JsonResponse({
            'success': False,
            'error': 'You have already checked in to this meeting.'
        })
    
    # Create attendance record
    attendance = MeetingAttendance.objects.create(
        meeting=meeting,
        executive=executive,
        check_in_method='qr_code',
        ip_address=get_client_ip(request),
        device_info=request.META.get('HTTP_USER_AGENT', '')[:200]
    )
    
    return JsonResponse({
        'success': True,
        'message': 'Check-in successful!',
        'check_in_time': attendance.check_in_time.strftime('%H:%M:%S')
    })


@login_required
@user_passes_test(lambda u: u.user_type == 'executive')
def meeting_attendance_list(request, pk):
    """View meeting attendance list"""
    meeting = get_object_or_404(ExecutiveMeeting, pk=pk)
    executive = get_object_or_404(StudentExecutive, user=request.user)
    
    # Only organizer can view full attendance
    if meeting.organized_by != executive:
        messages.error(request, 'Only the organizer can view attendance details.')
        return redirect('accounts:meeting_detail', pk=meeting.pk)
    
    attendance = MeetingAttendance.objects.filter(
        meeting=meeting
    ).select_related('executive__user').order_by('check_in_time')
    
    # Get all participants who haven't checked in
    checked_in_ids = attendance.values_list('executive_id', flat=True)
    absent = meeting.participants.exclude(id__in=checked_in_ids)
    
    context = {
        'meeting': meeting,
        'attendance': attendance,
        'absent': absent,
        'total_participants': meeting.participants.count(),
        'checked_in_count': attendance.count(),
    }
    
    return render(request, 'accounts/executive/meeting_attendance.html', context)

@login_required
def executive_meeting_create(request):
    """Create executive meeting"""
    
    if request.user.user_type not in ['executive', 'admin']:
        messages.error(request, "Access denied.")
        return redirect('core:home')

    if request.method == "POST":
        title = request.POST.get('title')
        description = request.POST.get('description')
        start_date = request.POST.get('start_date')

        Event.objects.create(
            title=title,
            description=description,
            start_date=start_date,
            created_by=request.user,
            is_active=True
        )

        messages.success(request, "Executive meeting created successfully.")
        return redirect('events:event_list')

    return render(request, 'core/executive_meeting_create.html')

@login_required
def executive_meeting_list(request):
    """List all executive meetings"""

    meetings = Event.objects.filter(
        is_active=True
    ).order_by('-start_date')

    context = {
        "meetings": meetings
    }

    return render(request, "accounts/executive/meeting_list.html", context)


@login_required
@user_passes_test(lambda u: u.user_type == 'executive')
def task_list(request):
    """List tasks"""
    executive = get_object_or_404(StudentExecutive, user=request.user)
    
    # Filter by status
    status = request.GET.get('status', 'pending')
    
    if status == 'assigned_by_me':
        tasks = ExecutiveTask.objects.filter(assigned_by=executive)
    elif status == 'completed':
        tasks = ExecutiveTask.objects.filter(
            assigned_to=executive,
            status='completed'
        )
    else:  # pending
        tasks = ExecutiveTask.objects.filter(
            assigned_to=executive
        ).exclude(status='completed')
    
    tasks = tasks.select_related('assigned_by', 'assigned_to', 'related_to_meeting')
    tasks = tasks.order_by('-priority', 'due_date')
    
    paginator = Paginator(tasks, 15)
    page = request.GET.get('page')
    tasks = paginator.get_page(page)
    
    context = {
        'tasks': tasks,
        'current_status': status,
        'executive': executive,
    }
    
    return render(request, 'accounts/executive/task_list.html', context)


@login_required
@user_passes_test(lambda u: u.user_type == 'executive')
def task_create(request):
    """Create new task"""
    executive = get_object_or_404(StudentExecutive, user=request.user)
    
    if request.method == 'POST':
        form = ExecutiveTaskForm(request.POST, request.FILES)
        if form.is_valid():
            task = form.save(commit=False)
            task.assigned_by = executive
            task.save()
            
            log_activity(
                request.user,
                'executive_action',
                request,
                {'action': 'task_created', 'task_id': task.id}
            )
            
            messages.success(request, f'Task "{task.title}" created successfully!')
            return redirect('accounts:task_list')
    else:
        # Pre-fill meeting if provided
        meeting_id = request.GET.get('meeting')
        if meeting_id:
            try:
                meeting = ExecutiveMeeting.objects.get(pk=meeting_id)
                form = ExecutiveTaskForm(initial={'related_to_meeting': meeting})
            except ExecutiveMeeting.DoesNotExist:
                form = ExecutiveTaskForm()
        else:
            form = ExecutiveTaskForm()
    
    return render(request, 'accounts/executive/task_form.html', {
        'form': form,
        'action': 'Create'
    })


@login_required
@user_passes_test(lambda u: u.user_type == 'executive')
def task_detail(request, pk):
    """View task details"""
    task = get_object_or_404(ExecutiveTask, pk=pk)
    executive = get_object_or_404(StudentExecutive, user=request.user)
    
    # Check permissions
    if task.assigned_to != executive and task.assigned_by != executive:
        messages.error(request, 'You do not have permission to view this task.')
        return redirect('accounts:task_list')
    
    if request.method == 'POST':
        form = ExecutiveTaskUpdateForm(request.POST, instance=task)
        if form.is_valid():
            if form.cleaned_data['status'] == 'completed' and task.status != 'completed':
                task.completed_date = timezone.now().date()
            
            form.save()
            
            messages.success(request, 'Task updated successfully!')
            return redirect('accounts:task_detail', pk=task.pk)
    else:
        form = ExecutiveTaskUpdateForm(instance=task)
    
    context = {
        'task': task,
        'form': form,
        'executive': executive,
        'can_edit': task.assigned_by == executive or task.assigned_to == executive,
    }
    
    return render(request, 'accounts/executive/task_detail.html', context)


@login_required
@user_passes_test(lambda u: u.user_type == 'executive')
def discussion_list(request):
    """List discussions"""
    executive = get_object_or_404(StudentExecutive, user=request.user)
    
    meeting_id = request.GET.get('meeting')
    if meeting_id:
        discussions = ExecutiveDiscussion.objects.filter(
            meeting_id=meeting_id
        ).select_related('created_by', 'meeting')
    else:
        discussions = ExecutiveDiscussion.objects.filter(
            Q(meeting__participants=executive) | Q(created_by=executive)
        ).distinct().select_related('created_by', 'meeting')
    
    discussions = discussions.order_by('-is_pinned', '-created_at')
    
    paginator = Paginator(discussions, 10)
    page = request.GET.get('page')
    discussions = paginator.get_page(page)
    
    context = {
        'discussions': discussions,
        'executive': executive,
    }
    
    return render(request, 'accounts/executive/discussion_list.html', context)


@login_required
@user_passes_test(lambda u: u.user_type == 'executive')
def discussion_create(request):
    """Create new discussion"""
    executive = get_object_or_404(StudentExecutive, user=request.user)
    
    if request.method == 'POST':
        form = ExecutiveDiscussionForm(request.POST, request.FILES)
        if form.is_valid():
            discussion = form.save(commit=False)
            discussion.created_by = executive
            
            # Set meeting if provided
            meeting_id = request.POST.get('meeting')
            if meeting_id:
                try:
                    discussion.meeting = ExecutiveMeeting.objects.get(pk=meeting_id)
                except ExecutiveMeeting.DoesNotExist:
                    pass
            
            discussion.save()
            
            messages.success(request, 'Discussion created successfully!')
            return redirect('accounts:discussion_detail', pk=discussion.pk)
    else:
        # Pre-fill meeting if provided
        meeting_id = request.GET.get('meeting')
        if meeting_id:
            try:
                meeting = ExecutiveMeeting.objects.get(pk=meeting_id)
                form = ExecutiveDiscussionForm(initial={'meeting': meeting})
            except ExecutiveMeeting.DoesNotExist:
                form = ExecutiveDiscussionForm()
        else:
            form = ExecutiveDiscussionForm()
    
    return render(request, 'accounts/executive/discussion_form.html', {
        'form': form,
        'action': 'Create'
    })


@login_required
@user_passes_test(lambda u: u.user_type == 'executive')
def discussion_detail(request, pk):
    """View discussion details"""
    discussion = get_object_or_404(ExecutiveDiscussion, pk=pk)
    executive = get_object_or_404(StudentExecutive, user=request.user)
    
    # Check permissions
    if discussion.meeting and executive not in discussion.meeting.participants.all():
        if discussion.created_by != executive:
            messages.error(request, 'You do not have permission to view this discussion.')
            return redirect('accounts:discussion_list')
    
    # Get comments
    comments = discussion.comments.select_related('author').order_by('created_at')
    
    # Comment form
    if request.method == 'POST':
        form = DiscussionCommentForm(request.POST, request.FILES)
        if form.is_valid():
            comment = form.save(commit=False)
            comment.discussion = discussion
            comment.author = executive
            comment.save()
            
            messages.success(request, 'Comment added successfully!')
            return redirect('accounts:discussion_detail', pk=discussion.pk)
    else:
        form = DiscussionCommentForm()
    
    context = {
        'discussion': discussion,
        'comments': comments,
        'form': form,
        'executive': executive,
        'can_edit': discussion.created_by == executive,
    }
    
    return render(request, 'accounts/executive/discussion_detail.html', context)


# ========== ADMIN MANAGEMENT VIEWS ==========

@login_required
@user_passes_test(lambda u: u.user_type == 'admin')
def user_list(request):
    """List all users (admin only)"""
    users = User.objects.all().select_related('account_executive_profile')
    
    # Apply filters
    form = UserSearchForm(request.GET)
    if form.is_valid():
        query = form.cleaned_data.get('query')
        user_type = form.cleaned_data.get('user_type')
        level = form.cleaned_data.get('level')
        program_type = form.cleaned_data.get('program_type')
        is_active = form.cleaned_data.get('is_active')
        
        if query:
            users = users.filter(
                Q(first_name__icontains=query) |
                Q(last_name__icontains=query) |
                Q(student_id__icontains=query) |
                Q(staff_id__icontains=query) |
                Q(email__icontains=query) |
                Q(username__icontains=query)
            )
        
        if user_type:
            users = users.filter(user_type=user_type)
        
        if level:
            users = users.filter(level=level)
        
        if program_type:
            users = users.filter(program_type=program_type)
        
        if is_active == 'true':
            users = users.filter(is_active=True)
        elif is_active == 'false':
            users = users.filter(is_active=False)
    
    # Pagination
    paginator = Paginator(users, 20)
    page = request.GET.get('page')
    users = paginator.get_page(page)
    
    # Statistics
    stats = {
        'total_students': User.objects.filter(user_type='student').count(),
        'total_staff': User.objects.filter(user_type='staff').count(),
        'total_executives': User.objects.filter(user_type='executive').count(),
        'total_admins': User.objects.filter(user_type='admin').count(),
        'active_users': User.objects.filter(is_active=True).count(),
        'inactive_users': User.objects.filter(is_active=False).count(),
    }
    
    context = {
        'users': users,
        'stats': stats,
        'form': form,
    }
    
    return render(request, 'accounts/admin/user_list.html', context)


@login_required
@user_passes_test(lambda u: u.user_type == 'admin')
def user_create(request):
    """Create new user (admin only)"""
    if request.method == 'POST':
        form = CustomUserCreationForm(request.POST, request.FILES)
        if form.is_valid():
            user = form.save()
            
            # Create executive profile if needed
            if user.user_type == 'executive':
                executive_position = request.POST.get('executive_position', 'member')
                StudentExecutive.objects.create(
                    user=user,
                    position=executive_position,
                    tenure_start_date=timezone.now().date(),
                    tenure_end_date=timezone.now().date() + timedelta(days=365),
                    appointed_by=request.user
                )
            
            # Log activity
            log_activity(
                request.user,
                'admin_action',
                request,
                {'action': 'user_created', 'user_id': user.id, 'username': user.username}
            )
            
            # Send password email if auto-generated
            if hasattr(form, 'temp_password'):
                send_password_email(user, form.temp_password)
                messages.success(
                    request, 
                    f'User {user.get_full_name()} created successfully! Password has been emailed.'
                )
            else:
                messages.success(request, f'User {user.get_full_name()} created successfully!')
            
            return redirect('accounts:user_detail', pk=user.pk)
    else:
        form = CustomUserCreationForm()
    
    return render(request, 'accounts/admin/user_form.html', {
        'form': form,
        'action': 'Create'
    })

# apps/accounts/views.py - Update user_detail function

@login_required
def user_detail(request, pk):
    """View user details"""
    user = get_object_or_404(User, pk=pk)
    
    # Check permissions
    if not (request.user.user_type == 'admin' or request.user.pk == user.pk):
        messages.error(request, 'You do not have permission to view this profile.')
        return redirect('core:home')
    
    context = {'profile_user': user}
    
    # Get user-specific data
    if user.user_type in ['student', 'executive']:
        from apps.courses.models import CourseRegistration
        from apps.payments.models import Payment
        from apps.core.models import AcademicSetting
        
        current_academic = AcademicSetting.objects.filter(is_active=True).first()
        
        # Registered courses
        registered_courses = CourseRegistration.objects.filter(
            student=user,
            academic_setting=current_academic
        ).select_related('course') if current_academic else []
        
        # Payment history
        payments = Payment.objects.filter(student=user).order_by('-created_at')
        
        # Payment statistics
        total_paid = payments.filter(status='success').aggregate(Sum('amount'))['amount__sum'] or 0
        pending_payments = payments.filter(status='pending').count()
        successful_payments = payments.filter(status='success').count()
        
        context.update({
            'registered_courses': registered_courses,
            'payments': payments[:10],  # Last 10 payments
            'total_paid': total_paid,
            'pending_payments': pending_payments,
            'successful_payments': successful_payments,
            'current_academic': current_academic,
        })
    
    if user.user_type == 'executive':
        from .models import StudentExecutive
        context['executive_profile'] = get_object_or_404(StudentExecutive, user=user)
    
    # Get recent activity with formatted details
    from .models import ActivityLog
    recent_activity = ActivityLog.objects.filter(
        user=user
    ).order_by('-timestamp')[:10]
    
    # Format activity details for display
    for activity in recent_activity:
        if activity.details:
            # Convert JSON to readable format
            if isinstance(activity.details, dict):
                details_list = []
                for key, value in activity.details.items():
                    if value:  # Only show non-empty values
                        # Format key to be more readable
                        formatted_key = key.replace('_', ' ').title()
                        details_list.append(f"{formatted_key}: {value}")
                activity.formatted_details = '<br>'.join(details_list)
            else:
                activity.formatted_details = str(activity.details)
        else:
            activity.formatted_details = ''
    
    context['recent_activity'] = recent_activity
    
    return render(request, 'accounts/admin/user_detail.html', context)

@login_required
@user_passes_test(lambda u: u.user_type == 'admin')
def user_edit(request, pk):
    """Edit user (admin only)"""
    user = get_object_or_404(User, pk=pk)
    
    if request.method == 'POST':
        form = CustomUserChangeForm(request.POST, request.FILES, instance=user)
        if form.is_valid():
            form.save()
            
            # Update executive profile if applicable
            if user.user_type == 'executive':
                executive, created = StudentExecutive.objects.get_or_create(user=user)
                executive.position = request.POST.get('executive_position', executive.position)
                executive.save()
            
            log_activity(
                request.user,
                'admin_action',
                request,
                {'action': 'user_updated', 'user_id': user.id}
            )
            
            messages.success(request, f'User {user.get_full_name()} updated successfully!')
            return redirect('accounts:user_detail', pk=user.pk)
    else:
        form = CustomUserChangeForm(instance=user)
    
    # Get executive position if applicable
    executive_position = None
    if user.user_type == 'executive':
        try:
            executive_position = user.executive_profile.position
        except StudentExecutive.DoesNotExist:
            pass
    
    return render(request, 'accounts/admin/user_form.html', {
        'form': form,
        'action': 'Edit',
        'user': user,
        'executive_position': executive_position
    })


@login_required
@user_passes_test(lambda u: u.user_type == 'admin')
@require_POST
def toggle_user_status(request, pk):
    """Toggle user active status (admin only)"""
    user = get_object_or_404(User, pk=pk)
    
    # Don't allow deactivating yourself
    if user == request.user:
        messages.error(request, 'You cannot deactivate your own account.')
        return redirect('accounts:user_detail', pk=user.pk)
    
    user.is_active = not user.is_active
    user.save()
    
    status = 'activated' if user.is_active else 'deactivated'
    
    log_activity(
        request.user,
        'admin_action',
        request,
        {'action': f'user_{status}', 'user_id': user.id}
    )
    
    messages.success(request, f'User {user.get_full_name()} {status} successfully!')
    
    return redirect('accounts:user_detail', pk=user.pk)


                            
                            # from datetime import date
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib import messages
from django.db import transaction
from django.shortcuts import render, redirect
import csv
import openpyxl  # Add this import for Excel files

from apps.accounts.models import User
from apps.accounts.forms import BulkStudentUploadForm


def generate_student_id(row):
    """
    Generates a new student ID based on year of admission.
    Format: BTMLYYNNNN (e.g., BTML260001)
    """
    # Get the year of admission
    year = row.get('year_of_admission')

    # Ensure year is an integer
    try:
        year = int(year)
    except (TypeError, ValueError):
        year = date.today().year  # fallback to current year

    # Prefix for this batch of students
    prefix = f'BTML{str(year)[-2:]}'  # e.g., BTML26

    # Get the last student with this prefix
    last_student = User.objects.filter(
        student_id__startswith=prefix
    ).order_by('-student_id').first()

    # Determine the new number
    if last_student:
        try:
            last_number = int(last_student.student_id[-4:])
        except ValueError:
            last_number = 0
        new_number = last_number + 1
    else:
        new_number = 1

    # Generate the new student ID
    student_id = f"{prefix}{new_number:04d}"

    return student_id


from datetime import date
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib import messages
from django.db import transaction
from django.shortcuts import render, redirect
import csv
import openpyxl

from apps.accounts.models import User
from apps.accounts.forms import BulkStudentUploadForm


@login_required
@user_passes_test(lambda u: u.user_type == 'admin')
def bulk_upload_students(request):
    """Bulk upload students via CSV/Excel (admin only)"""
    if request.method == 'POST':
        form = BulkStudentUploadForm(request.POST, request.FILES)
        if form.is_valid():
            file = request.FILES['file']
            created = 0
            errors = []
            
            with transaction.atomic():
                if file.name.endswith('.csv'):
                    # Process CSV
                    decoded_file = file.read().decode('utf-8').splitlines()
                    reader = csv.DictReader(decoded_file)
                    
                    for row_num, row in enumerate(reader, start=2):
                        try:
                            # Validate required fields
                            required_fields = ['first_name', 'last_name', 'year_of_admission', 'level']
                            missing = [f for f in required_fields if not row.get(f)]
                            if missing:
                                raise ValueError(f"Missing required fields: {', '.join(missing)}")
                            
                            # Create a temporary user object to generate the ID
                            temp_user = User(
                                year_of_admission=int(row['year_of_admission']),
                                first_name=row['first_name'].strip(),
                                last_name=row['last_name'].strip()
                            )
                            
                            # Generate student ID using the model method
                            student_id = temp_user.generate_student_id()
                            
                            # Create the actual user
                            user = User.objects.create_user(
                                username=student_id,
                                email=row.get('email', f"{student_id}@student.meltsa.edu"),
                                password='Student@123',
                                first_name=row['first_name'].strip(),
                                last_name=row['last_name'].strip(),
                                student_id=student_id,
                                year_of_admission=int(row['year_of_admission']),
                                level=int(row['level']),
                                program_type=row.get('program_type', 'regular'),
                                user_type='student',
                                requires_password_change=True
                            )
                            created += 1
                            
                        except Exception as e:
                            errors.append(f"Row {row_num}: {str(e)}")
                
                elif file.name.endswith(('.xlsx', '.xls')):
                    # Process Excel file
                    try:
                        wb = openpyxl.load_workbook(file)
                        ws = wb.active
                        
                        # Get headers from first row
                        headers = []
                        for cell in ws[1]:
                            headers.append(cell.value)
                        
                        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                            try:
                                # Create dictionary from row data
                                row_data = {}
                                for i, value in enumerate(row):
                                    if i < len(headers) and headers[i]:
                                        row_data[headers[i]] = value
                                
                                # Validate required fields
                                required_fields = ['first_name', 'last_name', 'year_of_admission', 'level']
                                missing = [f for f in required_fields if not row_data.get(f)]
                                if missing:
                                    raise ValueError(f"Missing required fields: {', '.join(missing)}")
                                
                                # Create a temporary user object to generate the ID
                                temp_user = User(
                                    year_of_admission=int(row_data['year_of_admission']),
                                    first_name=str(row_data['first_name']).strip(),
                                    last_name=str(row_data['last_name']).strip()
                                )
                                
                                # Generate student ID using the model method
                                student_id = temp_user.generate_student_id()
                                
                                # Create the actual user
                                user = User.objects.create_user(
                                    username=student_id,
                                    email=row_data.get('email', f"{student_id}@student.meltsa.edu"),
                                    password='Student@123',
                                    first_name=str(row_data['first_name']).strip(),
                                    last_name=str(row_data['last_name']).strip(),
                                    student_id=student_id,
                                    year_of_admission=int(row_data['year_of_admission']),
                                    level=int(row_data['year']),
                                    program_type=row_data.get('program_type', 'regular'),
                                    user_type='student',
                                    requires_password_change=True
                                )
                                created += 1
                                
                            except Exception as e:
                                errors.append(f"Row {row_num}: {str(e)}")
                                
                    except Exception as e:
                        errors.append(f"Error processing Excel file: {str(e)}")
            
            # Show results
            if created > 0:
                messages.success(request, f'Successfully created {created} students!')
            
            if errors:
                for error in errors[:5]:  # Show first 5 errors
                    messages.error(request, error)
                if len(errors) > 5:
                    messages.warning(request, f'And {len(errors) - 5} more errors...')
            
            return redirect('accounts:user_list')
    else:
        form = BulkStudentUploadForm()
    
    return render(request, 'accounts/admin/bulk_upload.html', {'form': form})


@login_required
@user_passes_test(lambda u: u.user_type == 'admin')
def export_users(request):
    """Export users to CSV (admin only)"""
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="meltsa_users.csv"'
    
    writer = csv.writer(response)
    writer.writerow([
        'Username', 'Student ID', 'Staff ID', 'First Name', 'Last Name', 'Email',
        'User Type', 'Level', 'Program Type', 'Year of Admission',
        'Is Active', 'Date Joined', 'Last Login', 'Phone Number'
    ])
    
    users = User.objects.all().order_by('date_joined')
    for user in users:
        writer.writerow([
            user.username,
            user.student_id or '',
            user.staff_id or '',
            user.first_name,
            user.last_name,
            user.email,
            user.user_type,
            user.level or '',
            user.program_type or '',
            user.year_of_admission or '',
            'Yes' if user.is_active else 'No',
            user.date_joined.strftime('%Y-%m-%d %H:%M:%S'),
            user.last_login.strftime('%Y-%m-%d %H:%M:%S') if user.last_login else '',
            user.phone_number,
        ])
    
    # Log activity
    log_activity(request.user, 'admin_action', request, {'action': 'export_users'})
    
    return response


@login_required
@user_passes_test(lambda u: u.user_type == 'admin')
def activity_logs(request):
    """View activity logs (admin only)"""
    logs = ActivityLog.objects.select_related('user').order_by('-timestamp')
    
    # Filters
    user_id = request.GET.get('user')
    if user_id:
        logs = logs.filter(user_id=user_id)
    
    action = request.GET.get('action')
    if action:
        logs = logs.filter(action_type=action)
    
    date_from = request.GET.get('from')
    if date_from:
        logs = logs.filter(timestamp__date__gte=date_from)
    
    date_to = request.GET.get('to')
    if date_to:
        logs = logs.filter(timestamp__date__lte=date_to)
    
    # Pagination
    paginator = Paginator(logs, 50)
    page = request.GET.get('page')
    logs = paginator.get_page(page)
    
    context = {
        'logs': logs,
        'users': User.objects.all(),
        'action_types': ActivityLog.ACTION_TYPES,
        'total_logs': ActivityLog.objects.count(),
        'today_logs': ActivityLog.objects.filter(timestamp__date=timezone.now().date()).count(),
    }
    
    return render(request, 'accounts/admin/activity_logs.html', context)


@login_required
@user_passes_test(lambda u: u.user_type == 'admin')
def export_activity_logs(request):
    """Export activity logs to CSV (admin only)"""
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="activity_logs.csv"'
    
    writer = csv.writer(response)
    writer.writerow(['User', 'Action', 'Timestamp', 'IP Address', 'Details'])
    
    logs = ActivityLog.objects.select_related('user').order_by('-timestamp')
    
    for log in logs:
        writer.writerow([
            log.user.get_full_name() or log.user.username,
            log.get_action_type_display(),
            log.timestamp.strftime('%Y-%m-%d %H:%M:%S'),
            log.ip_address,
            json.dumps(log.details),
        ])
    
    return response


@login_required
@user_passes_test(lambda u: u.user_type == 'admin')
def login_attempts(request):
    """View login attempts (admin only)"""
    attempts = LoginAttempt.objects.all().order_by('-timestamp')
    
    # Filters
    username = request.GET.get('username')
    if username:
        attempts = attempts.filter(username__icontains=username)
    
    ip = request.GET.get('ip')
    if ip:
        attempts = attempts.filter(ip_address__icontains=ip)
    
    successful = request.GET.get('successful')
    if successful == 'true':
        attempts = attempts.filter(successful=True)
    elif successful == 'false':
        attempts = attempts.filter(successful=False)
    
    # Pagination
    paginator = Paginator(attempts, 50)
    page = request.GET.get('page')
    attempts = paginator.get_page(page)
    
    context = {
        'attempts': attempts,
        'failed_count': LoginAttempt.objects.filter(successful=False).count(),
        'success_count': LoginAttempt.objects.filter(successful=True).count(),
    }
    
    return render(request, 'accounts/admin/login_attempts.html', context)






# apps/accounts/views.py (Add these to your existing views.py)

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib import messages
from django.db.models import Q, Count, Sum
from django.core.paginator import Paginator
from django.http import JsonResponse, HttpResponse
from django.utils import timezone
from django.views.decorators.http import require_POST
import csv
from datetime import timedelta

from .models import User, StudentExecutive, ExecutiveMeeting, ExecutiveTask, ExecutiveDiscussion, DiscussionComment, MeetingAttendance, ActivityLog
from .forms import (
    ExecutiveMeetingForm, ExecutiveTaskForm, ExecutiveTaskUpdateForm,
    ExecutiveDiscussionForm, DiscussionCommentForm
)


# ========== EXECUTIVE DASHBOARD ==========

@login_required
@user_passes_test(lambda u: u.user_type == 'executive')
def executive_dashboard(request):
    """Executive dashboard view"""
    executive = get_object_or_404(StudentExecutive, user=request.user)
    
    # Get upcoming meetings
    upcoming_meetings = ExecutiveMeeting.objects.filter(
        participants=executive,
        date__gte=timezone.now().date(),
        status='scheduled'
    ).order_by('date', 'start_time')[:5]
    
    # Get tasks assigned to executive
    my_tasks = ExecutiveTask.objects.filter(
        assigned_to=executive
    ).exclude(status='completed').order_by('due_date')[:5]
    
    # Get recent discussions
    recent_discussions = ExecutiveDiscussion.objects.filter(
        Q(created_by=executive) | Q(meeting__participants=executive)
    ).distinct().order_by('-created_at')[:5]
    
    # Get meetings organized by executive
    organized_meetings = ExecutiveMeeting.objects.filter(
        organized_by=executive,
        date__gte=timezone.now().date()
    ).order_by('date')[:3]
    
    # Get today's meeting if any
    todays_meeting = ExecutiveMeeting.objects.filter(
        organized_by=executive,
        date=timezone.now().date(),
        status='scheduled'
    ).first()
    
    # Get statistics
    stats = {
        'total_meetings': ExecutiveMeeting.objects.filter(
            participants=executive
        ).count(),
        'upcoming_count': upcoming_meetings.count(),
        'pending_tasks': my_tasks.count(),
        'attended_meetings': MeetingAttendance.objects.filter(
            executive=executive
        ).count(),
        'organized_meetings': ExecutiveMeeting.objects.filter(
            organized_by=executive
        ).count(),
        'discussions_count': ExecutiveDiscussion.objects.filter(
            created_by=executive
        ).count(),
    }
    
    # Get attendance rate
    total_invited = ExecutiveMeeting.objects.filter(
        participants=executive
    ).count()
    if total_invited > 0:
        stats['attendance_rate'] = (stats['attended_meetings'] / total_invited) * 100
    else:
        stats['attendance_rate'] = 0
    
    context = {
        'executive': executive,
        'upcoming_meetings': upcoming_meetings,
        'my_tasks': my_tasks,
        'recent_discussions': recent_discussions,
        'organized_meetings': organized_meetings,
        'todays_meeting': todays_meeting,
        'stats': stats,
    }
    
    return render(request, 'accounts/executive/dashboard.html', context)


# ========== MEETING MANAGEMENT ==========

@login_required
@user_passes_test(lambda u: u.user_type == 'executive')
def executive_meeting_list(request):
    """List all executive meetings"""
    executive = get_object_or_404(StudentExecutive, user=request.user)
    
    # Filter by status
    status = request.GET.get('status', 'upcoming')
    
    if status == 'upcoming':
        meetings = ExecutiveMeeting.objects.filter(
            participants=executive,
            date__gte=timezone.now().date(),
            status__in=['scheduled']
        )
    elif status == 'past':
        meetings = ExecutiveMeeting.objects.filter(
            participants=executive,
            date__lt=timezone.now().date()
        )
    elif status == 'organized':
        meetings = ExecutiveMeeting.objects.filter(
            organized_by=executive
        )
    else:
        meetings = ExecutiveMeeting.objects.filter(
            participants=executive
        )
    
    # Apply search
    search_query = request.GET.get('q')
    if search_query:
        meetings = meetings.filter(
            Q(title__icontains=search_query) |
            Q(description__icontains=search_query) |
            Q(venue__icontains=search_query)
        )
    
    # Apply type filter
    meeting_type = request.GET.get('type')
    if meeting_type:
        meetings = meetings.filter(meeting_type=meeting_type)
    
    # Apply date filters
    month = request.GET.get('month')
    year = request.GET.get('year')
    if month and year:
        meetings = meetings.filter(
            date__month=month,
            date__year=year
        )
    elif year:
        meetings = meetings.filter(date__year=year)
    
    meetings = meetings.order_by('-date', '-start_time')
    
    # Get attended meetings for status
    attended_meetings = MeetingAttendance.objects.filter(
        executive=executive
    ).values_list('meeting_id', flat=True)
    
    # Pagination
    paginator = Paginator(meetings, 12)
    page = request.GET.get('page')
    meetings = paginator.get_page(page)
    
    # Get meeting types for filter
    meeting_types = ExecutiveMeeting.MEETING_TYPES
    


    # Get years for filter
    years = ExecutiveMeeting.objects.dates('date', 'year')

    # Counts for tabs
    upcoming_count = ExecutiveMeeting.objects.filter(
        date__gte=timezone.now().date()
    ).count()
    
    # Counts for tabs
    upcoming_count = ExecutiveMeeting.objects.filter(
        participants=executive,
        date__gte=timezone.now().date()
    ).count()
    
    context = {
        'meetings': meetings,
        'current_status': status,
        'meeting_types': meeting_types,
        'years': years,
        'upcoming_count': upcoming_count,
        'attended_meetings': attended_meetings,
        'today': timezone.now().date(),
    }
    
    return render(request, 'accounts/executive/meeting_list.html', context)


@login_required
@user_passes_test(lambda u: u.user_type == 'executive')
def executive_meeting_create(request):
    """Create new executive meeting"""
    executive = get_object_or_404(StudentExecutive, user=request.user)
    
    if request.method == 'POST':
        form = ExecutiveMeetingForm(request.POST, request.FILES)
        if form.is_valid():
            meeting = form.save(commit=False)
            meeting.organized_by = executive
            meeting.save()
            form.save_m2m()  # Save participants
            
            # Add organizer as participant if not already included
            if executive not in meeting.participants.all():
                meeting.participants.add(executive)
            
            # Generate QR code
            meeting.generate_qr_code()
            meeting.save()
            
            # Log activity
            ActivityLog.objects.create(
                user=request.user,
                action_type='executive_action',
                ip_address=get_client_ip(request),
                details={'action': 'meeting_created', 'meeting_id': meeting.id, 'title': meeting.title}
            )
            
            messages.success(request, f'Meeting "{meeting.title}" created successfully!')
            return redirect('accounts:executive_meeting_detail', pk=meeting.pk)
    else:
        form = ExecutiveMeetingForm()
    
    # Get recent meetings for sidebar
    recent_meetings = ExecutiveMeeting.objects.filter(
        organized_by=executive
    ).order_by('-date')[:5]
    
    # Get all executives for participants field
    executives = StudentExecutive.objects.filter(tenure_status='active')
    
    context = {
        'form': form,
        'action': 'Create',
        'recent_meetings': recent_meetings,
        'executives': executives,
    }
    
    return render(request, 'accounts/executive/meeting_form.html', context)


@login_required
@user_passes_test(lambda u: u.user_type == 'executive')
def executive_meeting_detail(request, pk):
    """View meeting details"""
    meeting = get_object_or_404(ExecutiveMeeting, pk=pk)
    executive = get_object_or_404(StudentExecutive, user=request.user)
    
    # Check if user can view this meeting
    if executive not in meeting.participants.all() and meeting.organized_by != executive:
        messages.error(request, 'You do not have permission to view this meeting.')
        return redirect('accounts:executive_meeting_list')
    
    # Get attendance
    attendance = MeetingAttendance.objects.filter(meeting=meeting).select_related('executive__user')
    
    # Check if user has checked in
    has_checked_in = attendance.filter(executive=executive).exists()
    
    # Get discussions related to this meeting
    discussions = ExecutiveDiscussion.objects.filter(meeting=meeting).order_by('-created_at')
    
    # Get tasks related to meeting
    tasks = ExecutiveTask.objects.filter(related_to_meeting=meeting)
    
    # Get attendance statistics
    total_participants = meeting.participants.count()
    checked_in_count = attendance.count()
    
    context = {
        'meeting': meeting,
        'attendance': attendance,
        'has_checked_in': has_checked_in,
        'discussions': discussions,
        'tasks': tasks,
        'executive': executive,
        'is_organizer': meeting.organized_by == executive,
        'total_participants': total_participants,
        'checked_in_count': checked_in_count,
        'attendance_rate': (checked_in_count / total_participants * 100) if total_participants > 0 else 0,
    }
    
    return render(request, 'accounts/executive/meeting_detail.html', context)


@login_required
@user_passes_test(lambda u: u.user_type == 'executive')
def executive_meeting_edit(request, pk):
    """Edit meeting"""
    meeting = get_object_or_404(ExecutiveMeeting, pk=pk)
    executive = get_object_or_404(StudentExecutive, user=request.user)
    
    # Only organizer can edit
    if meeting.organized_by != executive:
        messages.error(request, 'Only the meeting organizer can edit this meeting.')
        return redirect('accounts:executive_meeting_detail', pk=meeting.pk)
    
    if request.method == 'POST':
        form = ExecutiveMeetingForm(request.POST, request.FILES, instance=meeting)
        if form.is_valid():
            form.save()
            messages.success(request, 'Meeting updated successfully!')
            return redirect('accounts:executive_meeting_detail', pk=meeting.pk)
    else:
        form = ExecutiveMeetingForm(instance=meeting)
    
    context = {
        'form': form,
        'action': 'Edit',
        'meeting': meeting,
    }
    
    return render(request, 'accounts/executive/meeting_form.html', context)


@login_required
@require_POST
@user_passes_test(lambda u: u.user_type == 'executive')
def executive_meeting_delete(request, pk):
    """Delete meeting"""
    meeting = get_object_or_404(ExecutiveMeeting, pk=pk)
    executive = get_object_or_404(StudentExecutive, user=request.user)
    
    if meeting.organized_by != executive:
        return JsonResponse({'success': False, 'error': 'Permission denied'}, status=403)
    
    meeting.delete()
    
    ActivityLog.objects.create(
        user=request.user,
        action_type='executive_action',
        ip_address=get_client_ip(request),
        details={'action': 'meeting_deleted', 'meeting_id': pk}
    )
    
    return JsonResponse({'success': True})


@login_required
@user_passes_test(lambda u: u.user_type == 'executive')
def executive_meeting_attendance(request, pk):
    """View and manage meeting attendance"""
    meeting = get_object_or_404(ExecutiveMeeting, pk=pk)
    executive = get_object_or_404(StudentExecutive, user=request.user)
    
    # Only organizer can view full attendance
    if meeting.organized_by != executive:
        messages.error(request, 'Only the organizer can view attendance details.')
        return redirect('accounts:executive_meeting_detail', pk=meeting.pk)
    
    # Get attendance records
    attendance = MeetingAttendance.objects.filter(
        meeting=meeting
    ).select_related('executive__user').order_by('check_in_time')
    
    # Get all participants who haven't checked in
    checked_in_ids = attendance.values_list('executive_id', flat=True)
    absent = meeting.participants.exclude(id__in=checked_in_ids)
    
    # Handle manual check-in via POST
    if request.method == 'POST':
        executive_id = request.POST.get('executive_id')
        try:
            attendee = StudentExecutive.objects.get(id=executive_id)
            if attendee in meeting.participants.all():
                MeetingAttendance.objects.create(
                    meeting=meeting,
                    executive=attendee,
                    check_in_method='manual',
                    ip_address=get_client_ip(request)
                )
                messages.success(request, f'Checked in {attendee.user.get_full_name()}')
            else:
                messages.error(request, 'Selected executive is not a participant')
        except StudentExecutive.DoesNotExist:
            messages.error(request, 'Executive not found')
        
        return redirect('accounts:executive_meeting_attendance', pk=meeting.pk)
    
    context = {
        'meeting': meeting,
        'attendance': attendance,
        'absent': absent,
        'total_participants': meeting.participants.count(),
        'checked_in_count': attendance.count(),
    }
    
    return render(request, 'accounts/executive/meeting_attendance.html', context)


@login_required
@require_POST
@user_passes_test(lambda u: u.user_type == 'executive')
def executive_meeting_check_in(request, pk):
    """Check in to meeting via QR code or code"""
    meeting = get_object_or_404(ExecutiveMeeting, pk=pk)
    executive = get_object_or_404(StudentExecutive, user=request.user)
    
    # Check if meeting is today
    if meeting.date != timezone.now().date():
        return JsonResponse({
            'success': False,
            'error': 'This meeting is not scheduled for today.'
        })
    
    # Check if already checked in
    if MeetingAttendance.objects.filter(meeting=meeting, executive=executive).exists():
        return JsonResponse({
            'success': False,
            'error': 'You have already checked in to this meeting.'
        })
    
    # Check if user is a participant
    if executive not in meeting.participants.all():
        return JsonResponse({
            'success': False,
            'error': 'You are not a participant in this meeting.'
        })
    
    # Create attendance record
    attendance = MeetingAttendance.objects.create(
        meeting=meeting,
        executive=executive,
        check_in_method='qr_code',
        ip_address=get_client_ip(request),
        device_info=request.META.get('HTTP_USER_AGENT', '')[:200]
    )
    
    return JsonResponse({
        'success': True,
        'message': 'Check-in successful!',
        'check_in_time': attendance.check_in_time.strftime('%H:%M:%S')
    })


@login_required
@user_passes_test(lambda u: u.user_type == 'executive')
def export_meeting_attendance(request, pk):
    """Export meeting attendance to CSV"""
    meeting = get_object_or_404(ExecutiveMeeting, pk=pk)
    executive = get_object_or_404(StudentExecutive, user=request.user)
    
    if meeting.organized_by != executive:
        messages.error(request, 'Permission denied')
        return redirect('accounts:executive_meeting_detail', pk=meeting.pk)
    
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="meeting_{meeting.id}_attendance.csv"'
    
    writer = csv.writer(response)
    writer.writerow(['Name', 'Position', 'Check-in Time', 'Method', 'IP Address'])
    
    attendance = MeetingAttendance.objects.filter(meeting=meeting).select_related('executive__user')
    for record in attendance:
        writer.writerow([
            record.executive.user.get_full_name(),
            record.executive.get_position_display(),
            record.check_in_time.strftime('%Y-%m-%d %H:%M:%S'),
            record.get_check_in_method_display(),
            record.ip_address or 'N/A'
        ])
    
    return response


@login_required
@require_POST
@user_passes_test(lambda u: u.user_type == 'executive')
def send_meeting_reminders(request, pk):
    """Send reminders to participants"""
    meeting = get_object_or_404(ExecutiveMeeting, pk=pk)
    executive = get_object_or_404(StudentExecutive, user=request.user)
    
    if meeting.organized_by != executive:
        return JsonResponse({'success': False, 'error': 'Permission denied'}, status=403)
    
    # Here you would implement email/SMS sending logic
    # For now, just log and return success
    
    ActivityLog.objects.create(
        user=request.user,
        action_type='executive_action',
        ip_address=get_client_ip(request),
        details={'action': 'reminders_sent', 'meeting_id': meeting.id}
    )
    
    return JsonResponse({'success': True, 'message': 'Reminders sent successfully'})


@login_required
@user_passes_test(lambda u: u.user_type == 'executive')
def export_all_meetings(request):
    """Export all meetings to CSV"""
    executive = get_object_or_404(StudentExecutive, user=request.user)
    
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="all_meetings.csv"'
    
    writer = csv.writer(response)
    writer.writerow(['Title', 'Type', 'Date', 'Time', 'Venue', 'Status', 'Participants', 'Attendance Rate'])
    
    meetings = ExecutiveMeeting.objects.filter(participants=executive)
    for meeting in meetings:
        total_participants = meeting.participants.count()
        attendance_count = MeetingAttendance.objects.filter(meeting=meeting).count()
        attendance_rate = (attendance_count / total_participants * 100) if total_participants > 0 else 0
        
        writer.writerow([
            meeting.title,
            meeting.get_meeting_type_display(),
            meeting.date.strftime('%Y-%m-%d'),
            f"{meeting.start_time.strftime('%H:%M')} - {meeting.end_time.strftime('%H:%M')}",
            meeting.venue,
            meeting.get_status_display(),
            total_participants,
            f"{attendance_rate:.1f}%"
        ])
    
    return response


# ========== TASK MANAGEMENT ==========

@login_required
@user_passes_test(lambda u: u.user_type == 'executive')
def executive_task_list(request):
    """List tasks"""
    executive = get_object_or_404(StudentExecutive, user=request.user)
    
    # Filter by status
    status = request.GET.get('status', 'pending')
    
    if status == 'assigned_by_me':
        tasks = ExecutiveTask.objects.filter(assigned_by=executive)
    elif status == 'completed':
        tasks = ExecutiveTask.objects.filter(
            assigned_to=executive,
            status='completed'
        )
    else:  # pending
        tasks = ExecutiveTask.objects.filter(
            assigned_to=executive
        ).exclude(status='completed')
    
    # Apply search
    search_query = request.GET.get('q')
    if search_query:
        tasks = tasks.filter(
            Q(title__icontains=search_query) |
            Q(description__icontains=search_query)
        )
    
    tasks = tasks.select_related('assigned_by', 'assigned_to', 'related_to_meeting')
    tasks = tasks.order_by('-priority', 'due_date')
    
    paginator = Paginator(tasks, 15)
    page = request.GET.get('page')
    tasks = paginator.get_page(page)
    
    # Statistics
    stats = {
        'total': ExecutiveTask.objects.filter(assigned_to=executive).count(),
        'pending': ExecutiveTask.objects.filter(assigned_to=executive, status='pending').count(),
        'in_progress': ExecutiveTask.objects.filter(assigned_to=executive, status='in_progress').count(),
        'completed': ExecutiveTask.objects.filter(assigned_to=executive, status='completed').count(),
        'overdue': ExecutiveTask.objects.filter(
            assigned_to=executive,
            due_date__lt=timezone.now().date(),
            status__in=['pending', 'in_progress']
        ).count(),
    }
    
    context = {
        'tasks': tasks,
        'current_status': status,
        'executive': executive,
        'stats': stats,
    }
    
    return render(request, 'accounts/executive/task_list.html', context)


@login_required
@user_passes_test(lambda u: u.user_type == 'executive')
def executive_task_create(request):
    """Create new task"""
    executive = get_object_or_404(StudentExecutive, user=request.user)
    
    if request.method == 'POST':
        form = ExecutiveTaskForm(request.POST, request.FILES)
        if form.is_valid():
            task = form.save(commit=False)
            task.assigned_by = executive
            task.save()
            
            ActivityLog.objects.create(
                user=request.user,
                action_type='executive_action',
                ip_address=get_client_ip(request),
                details={'action': 'task_created', 'task_id': task.id, 'title': task.title}
            )
            
            messages.success(request, f'Task "{task.title}" created successfully!')
            return redirect('accounts:executive_task_detail', pk=task.pk)
    else:
        # Pre-fill meeting if provided
        meeting_id = request.GET.get('meeting')
        if meeting_id:
            try:
                meeting = ExecutiveMeeting.objects.get(pk=meeting_id)
                form = ExecutiveTaskForm(initial={'related_to_meeting': meeting})
            except ExecutiveMeeting.DoesNotExist:
                form = ExecutiveTaskForm()
        else:
            form = ExecutiveTaskForm()
    
    context = {
        'form': form,
        'action': 'Create',
    }
    
    return render(request, 'accounts/executive/task_form.html', context)


@login_required
@user_passes_test(lambda u: u.user_type == 'executive')
def executive_task_detail(request, pk):
    """View task details"""
    task = get_object_or_404(ExecutiveTask, pk=pk)
    executive = get_object_or_404(StudentExecutive, user=request.user)
    
    # Check permissions
    if task.assigned_to != executive and task.assigned_by != executive:
        messages.error(request, 'You do not have permission to view this task.')
        return redirect('accounts:executive_task_list')
    
    if request.method == 'POST':
        form = ExecutiveTaskUpdateForm(request.POST, instance=task)
        if form.is_valid():
            if form.cleaned_data['status'] == 'completed' and task.status != 'completed':
                task.completed_date = timezone.now().date()
            
            form.save()
            
            messages.success(request, 'Task updated successfully!')
            return redirect('accounts:executive_task_detail', pk=task.pk)
    else:
        form = ExecutiveTaskUpdateForm(instance=task)
    
    context = {
        'task': task,
        'form': form,
        'executive': executive,
        'can_edit': task.assigned_by == executive or task.assigned_to == executive,
    }
    
    return render(request, 'accounts/executive/task_detail.html', context)


@login_required
@user_passes_test(lambda u: u.user_type == 'executive')
def executive_task_edit(request, pk):
    """Edit task"""
    task = get_object_or_404(ExecutiveTask, pk=pk)
    executive = get_object_or_404(StudentExecutive, user=request.user)
    
    if task.assigned_by != executive:
        messages.error(request, 'Only the creator can edit this task.')
        return redirect('accounts:executive_task_detail', pk=task.pk)
    
    if request.method == 'POST':
        form = ExecutiveTaskForm(request.POST, request.FILES, instance=task)
        if form.is_valid():
            form.save()
            messages.success(request, 'Task updated successfully!')
            return redirect('accounts:executive_task_detail', pk=task.pk)
    else:
        form = ExecutiveTaskForm(instance=task)
    
    context = {
        'form': form,
        'action': 'Edit',
        'task': task,
    }
    
    return render(request, 'accounts/executive/task_form.html', context)


@login_required
@require_POST
@user_passes_test(lambda u: u.user_type == 'executive')
def executive_task_delete(request, pk):
    """Delete task"""
    task = get_object_or_404(ExecutiveTask, pk=pk)
    executive = get_object_or_404(StudentExecutive, user=request.user)
    
    if task.assigned_by != executive:
        return JsonResponse({'success': False, 'error': 'Permission denied'}, status=403)
    
    task.delete()
    return JsonResponse({'success': True})


@login_required
@require_POST
@user_passes_test(lambda u: u.user_type == 'executive')
def executive_task_update_status(request, pk):
    """Update task status via AJAX"""
    task = get_object_or_404(ExecutiveTask, pk=pk)
    executive = get_object_or_404(StudentExecutive, user=request.user)
    
    if task.assigned_to != executive:
        return JsonResponse({'success': False, 'error': 'Permission denied'}, status=403)
    
    status = request.POST.get('status')
    if status in dict(ExecutiveTask.TASK_STATUS):
        task.status = status
        if status == 'completed':
            task.completed_date = timezone.now().date()
        task.save()
        return JsonResponse({'success': True})
    
    return JsonResponse({'success': False, 'error': 'Invalid status'})


@login_required
@require_POST
@user_passes_test(lambda u: u.user_type == 'executive')
def executive_task_assign(request, pk):
    """Reassign task"""
    task = get_object_or_404(ExecutiveTask, pk=pk)
    executive = get_object_or_404(StudentExecutive, user=request.user)
    
    if task.assigned_by != executive:
        return JsonResponse({'success': False, 'error': 'Permission denied'}, status=403)
    
    assign_to_id = request.POST.get('assign_to')
    try:
        new_assignee = StudentExecutive.objects.get(id=assign_to_id)
        task.assigned_to = new_assignee
        task.save()
        return JsonResponse({'success': True})
    except StudentExecutive.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'User not found'})


@login_required
@user_passes_test(lambda u: u.user_type == 'executive')
def export_tasks(request):
    """Export tasks to CSV"""
    executive = get_object_or_404(StudentExecutive, user=request.user)
    
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="tasks.csv"'
    
    writer = csv.writer(response)
    writer.writerow(['Title', 'Description', 'Assigned To', 'Assigned By', 'Priority', 'Status', 'Due Date', 'Completed Date'])
    
    tasks = ExecutiveTask.objects.filter(Q(assigned_to=executive) | Q(assigned_by=executive))
    for task in tasks:
        writer.writerow([
            task.title,
            task.description,
            task.assigned_to.user.get_full_name() if task.assigned_to else 'Unassigned',
            task.assigned_by.user.get_full_name() if task.assigned_by else 'Unknown',
            task.get_priority_display(),
            task.get_status_display(),
            task.due_date.strftime('%Y-%m-%d') if task.due_date else '',
            task.completed_date.strftime('%Y-%m-%d') if task.completed_date else ''
        ])
    
    return response


# ========== DISCUSSION FORUM ==========

@login_required
@user_passes_test(lambda u: u.user_type == 'executive')
def executive_discussion_list(request):
    """List discussions"""
    executive = get_object_or_404(StudentExecutive, user=request.user)
    
    meeting_id = request.GET.get('meeting')
    if meeting_id:
        discussions = ExecutiveDiscussion.objects.filter(
            meeting_id=meeting_id
        ).select_related('created_by', 'meeting')
    else:
        discussions = ExecutiveDiscussion.objects.filter(
            Q(meeting__participants=executive) | Q(created_by=executive)
        ).distinct().select_related('created_by', 'meeting')
    
    # Apply search
    search_query = request.GET.get('q')
    if search_query:
        discussions = discussions.filter(
            Q(title__icontains=search_query) |
            Q(content__icontains=search_query)
        )
    
    discussions = discussions.order_by('-is_pinned', '-created_at')
    
    paginator = Paginator(discussions, 10)
    page = request.GET.get('page')
    discussions = paginator.get_page(page)
    
    context = {
        'discussions': discussions,
        'executive': executive,
    }
    
    return render(request, 'accounts/executive/discussion_list.html', context)


@login_required
@user_passes_test(lambda u: u.user_type == 'executive')
def executive_discussion_create(request):
    """Create new discussion"""
    executive = get_object_or_404(StudentExecutive, user=request.user)
    
    if request.method == 'POST':
        form = ExecutiveDiscussionForm(request.POST, request.FILES)
        if form.is_valid():
            discussion = form.save(commit=False)
            discussion.created_by = executive
            
            # Set meeting if provided
            meeting_id = request.POST.get('meeting')
            if meeting_id:
                try:
                    discussion.meeting = ExecutiveMeeting.objects.get(pk=meeting_id)
                except ExecutiveMeeting.DoesNotExist:
                    pass
            
            discussion.save()
            
            ActivityLog.objects.create(
                user=request.user,
                action_type='executive_action',
                ip_address=get_client_ip(request),
                details={'action': 'discussion_created', 'discussion_id': discussion.id, 'title': discussion.title}
            )
            
            messages.success(request, 'Discussion created successfully!')
            return redirect('accounts:executive_discussion_detail', pk=discussion.pk)
    else:
        # Pre-fill meeting if provided
        meeting_id = request.GET.get('meeting')
        if meeting_id:
            try:
                meeting = ExecutiveMeeting.objects.get(pk=meeting_id)
                form = ExecutiveDiscussionForm(initial={'meeting': meeting})
            except ExecutiveMeeting.DoesNotExist:
                form = ExecutiveDiscussionForm()
        else:
            form = ExecutiveDiscussionForm()
    
    context = {
        'form': form,
        'action': 'Create',
    }
    
    return render(request, 'accounts/executive/discussion_form.html', context)


@login_required
@user_passes_test(lambda u: u.user_type == 'executive')
def executive_discussion_detail(request, pk):
    """View discussion details"""
    discussion = get_object_or_404(ExecutiveDiscussion, pk=pk)
    executive = get_object_or_404(StudentExecutive, user=request.user)
    
    # Check permissions
    if discussion.meeting and executive not in discussion.meeting.participants.all():
        if discussion.created_by != executive:
            messages.error(request, 'You do not have permission to view this discussion.')
            return redirect('accounts:executive_discussion_list')
    
    # Get comments
    comments = discussion.comments.select_related('author__user').order_by('created_at')
    
    # Comment form
    if request.method == 'POST':
        form = DiscussionCommentForm(request.POST, request.FILES)
        if form.is_valid():
            comment = form.save(commit=False)
            comment.discussion = discussion
            comment.author = executive
            comment.save()
            
            messages.success(request, 'Comment added successfully!')
            return redirect('accounts:executive_discussion_detail', pk=discussion.pk)
    else:
        form = DiscussionCommentForm()
    
    context = {
        'discussion': discussion,
        'comments': comments,
        'form': form,
        'executive': executive,
        'can_edit': discussion.created_by == executive,
    }
    
    return render(request, 'accounts/executive/discussion_detail.html', context)


@login_required
@user_passes_test(lambda u: u.user_type == 'executive')
def executive_discussion_edit(request, pk):
    """Edit discussion"""
    discussion = get_object_or_404(ExecutiveDiscussion, pk=pk)
    executive = get_object_or_404(StudentExecutive, user=request.user)
    
    if discussion.created_by != executive:
        messages.error(request, 'You can only edit your own discussions.')
        return redirect('accounts:executive_discussion_detail', pk=discussion.pk)
    
    if request.method == 'POST':
        form = ExecutiveDiscussionForm(request.POST, request.FILES, instance=discussion)
        if form.is_valid():
            form.save()
            messages.success(request, 'Discussion updated successfully!')
            return redirect('accounts:executive_discussion_detail', pk=discussion.pk)
    else:
        form = ExecutiveDiscussionForm(instance=discussion)
    
    context = {
        'form': form,
        'action': 'Edit',
        'discussion': discussion,
    }
    
    return render(request, 'accounts/executive/discussion_form.html', context)


@login_required
@require_POST
@user_passes_test(lambda u: u.user_type == 'executive')
def executive_discussion_delete(request, pk):
    """Delete discussion"""
    discussion = get_object_or_404(ExecutiveDiscussion, pk=pk)
    executive = get_object_or_404(StudentExecutive, user=request.user)
    
    if discussion.created_by != executive:
        return JsonResponse({'success': False, 'error': 'Permission denied'}, status=403)
    
    discussion.delete()
    return JsonResponse({'success': True})


@login_required
@require_POST
@user_passes_test(lambda u: u.user_type == 'executive')
def executive_discussion_pin(request, pk):
    """Pin/unpin discussion"""
    discussion = get_object_or_404(ExecutiveDiscussion, pk=pk)
    executive = get_object_or_404(StudentExecutive, user=request.user)
    
    # Only admins or meeting organizers can pin
    if not (request.user.user_type == 'admin' or 
            (discussion.meeting and discussion.meeting.organized_by == executive)):
        return JsonResponse({'success': False, 'error': 'Permission denied'}, status=403)
    
    discussion.is_pinned = not discussion.is_pinned
    discussion.save()
    
    return JsonResponse({'success': True, 'is_pinned': discussion.is_pinned})


@login_required
@require_POST
@user_passes_test(lambda u: u.user_type == 'executive')
def executive_discussion_add_comment(request, pk):
    """Add comment to discussion (AJAX)"""
    discussion = get_object_or_404(ExecutiveDiscussion, pk=pk)
    executive = get_object_or_404(StudentExecutive, user=request.user)
    
    content = request.POST.get('content')
    if not content:
        return JsonResponse({'success': False, 'error': 'Content is required'})
    
    comment = DiscussionComment.objects.create(
        discussion=discussion,
        author=executive,
        content=content
    )
    
    return JsonResponse({
        'success': True,
        'comment': {
            'id': comment.id,
            'author': comment.author.user.get_full_name(),
            'content': comment.content,
            'created_at': comment.created_at.strftime('%Y-%m-%d %H:%M'),
        }
    })


@login_required
@require_POST
@user_passes_test(lambda u: u.user_type == 'executive')
def executive_discussion_comment_delete(request, pk):
    """Delete comment"""
    comment = get_object_or_404(DiscussionComment, pk=pk)
    executive = get_object_or_404(StudentExecutive, user=request.user)
    
    if comment.author != executive:
        return JsonResponse({'success': False, 'error': 'Permission denied'}, status=403)
    
    comment.delete()
    return JsonResponse({'success': True})


@login_required
@require_POST
@user_passes_test(lambda u: u.user_type == 'executive')
def executive_discussion_comment_edit(request, pk):
    """Edit comment"""
    comment = get_object_or_404(DiscussionComment, pk=pk)
    executive = get_object_or_404(StudentExecutive, user=request.user)
    
    if comment.author != executive:
        return JsonResponse({'success': False, 'error': 'Permission denied'}, status=403)
    
    content = request.POST.get('content')
    if not content:
        return JsonResponse({'success': False, 'error': 'Content is required'})
    
    comment.content = content
    comment.save()
    
    return JsonResponse({'success': True})


@login_required
@require_POST
@user_passes_test(lambda u: u.user_type == 'executive')
def executive_discussion_like(request, pk):
    """Like/unlike discussion"""
    discussion = get_object_or_404(ExecutiveDiscussion, pk=pk)
    executive = get_object_or_404(StudentExecutive, user=request.user)
    
    # This would require a DiscussionLike model
    # Placeholder for now
    return JsonResponse({'success': True})


@login_required
@require_POST
@user_passes_test(lambda u: u.user_type == 'executive')
def executive_discussion_comment_like(request, pk):
    """Like/unlike comment"""
    comment = get_object_or_404(DiscussionComment, pk=pk)
    executive = get_object_or_404(StudentExecutive, user=request.user)
    
    # This would require a CommentLike model
    # Placeholder for now
    return JsonResponse({'success': True})

# apps/accounts/views.py - Add these functions

@login_required
@user_passes_test(lambda u: u.user_type == 'admin')
def reset_user_password(request, pk):
    """Reset user password (admin only)"""
    if request.method == 'POST':
        user = get_object_or_404(User, pk=pk)
        
        new_password = request.POST.get('new_password')
        confirm_password = request.POST.get('confirm_password')
        require_change = request.POST.get('require_change') == 'on'
        
        if new_password != confirm_password:
            messages.error(request, 'Passwords do not match.')
            return redirect('accounts:user_detail', pk=pk)
        
        if len(new_password) < 8:
            messages.error(request, 'Password must be at least 8 characters long.')
            return redirect('accounts:user_detail', pk=pk)
        
        user.set_password(new_password)
        user.requires_password_change = require_change
        user.save()
        
        # Send email notification
        # send_password_reset_email(user, new_password)
        
        messages.success(request, f'Password reset successfully for {user.get_full_name()}')
        return redirect('accounts:user_detail', pk=pk)
    
    return redirect('accounts:user_detail', pk=pk)


@login_required
@user_passes_test(lambda u: u.user_type == 'admin')
def unlock_user_account(request, pk):
    """Unlock user account (admin only)"""
    if request.method == 'POST':
        user = get_object_or_404(User, pk=pk)
        
        user.failed_login_attempts = 0
        user.account_locked_until = None
        user.save()
        
        messages.success(request, f'Account unlocked successfully for {user.get_full_name()}')
        return redirect('accounts:user_detail', pk=pk)
    
    return redirect('accounts:user_detail', pk=pk)



# apps/accounts/views.py - Add these functions

from .models import ExecutiveTask, TaskComment  # You need to create this model

@login_required
@user_passes_test(lambda u: u.user_type == 'executive')
def executive_task_add_comment(request, pk):
    """Add a comment to a task"""
    task = get_object_or_404(ExecutiveTask, pk=pk)
    executive = get_object_or_404(StudentExecutive, user=request.user)
    
    # Check if user has permission to comment (assigned to or assigned by)
    if task.assigned_to != executive and task.assigned_by != executive:
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'success': False, 'error': 'Permission denied'})
        messages.error(request, 'You do not have permission to comment on this task.')
        return redirect('accounts:executive_task_detail', pk=pk)
    
    if request.method == 'POST':
        content = request.POST.get('content')
        
        if not content:
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({'success': False, 'error': 'Comment content is required'})
            messages.error(request, 'Comment cannot be empty.')
            return redirect('accounts:executive_task_detail', pk=pk)
        
        # Create comment - you need a TaskComment model
        comment = TaskComment.objects.create(
            task=task,
            author=executive,
            content=content
        )
        
        # Log activity
        ActivityLog.objects.create(
            user=request.user,
            action_type='executive_action',
            ip_address=get_client_ip(request),
            details={'action': 'task_comment_added', 'task_id': task.id}
        )
        
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({
                'success': True,
                'comment': {
                    'id': comment.id,
                    'author': comment.author.user.get_full_name(),
                    'content': comment.content,
                    'created_at': comment.created_at.strftime('%Y-%m-%d %H:%M:%S')
                }
            })
        
        messages.success(request, 'Comment added successfully.')
        return redirect('accounts:executive_task_detail', pk=pk)
    
    return redirect('accounts:executive_task_detail', pk=pk)


@login_required
@user_passes_test(lambda u: u.user_type == 'executive')
def executive_task_comment_delete(request, pk):
    """Delete a task comment"""
    comment = get_object_or_404(TaskComment, pk=pk)
    executive = get_object_or_404(StudentExecutive, user=request.user)
    
    # Check if user is the author
    if comment.author != executive:
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'success': False, 'error': 'Permission denied'})
        messages.error(request, 'You can only delete your own comments.')
        return redirect('accounts:executive_task_detail', pk=comment.task.pk)
    
    task_id = comment.task.id
    comment.delete()
    
    # Log activity
    ActivityLog.objects.create(
        user=request.user,
        action_type='executive_action',
        ip_address=get_client_ip(request),
        details={'action': 'task_comment_deleted', 'task_id': task_id}
    )
    
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return JsonResponse({'success': True})
    
    messages.success(request, 'Comment deleted successfully.')
    return redirect('accounts:executive_task_detail', pk=task_id)